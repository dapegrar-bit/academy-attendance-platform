import csv
import calendar
from collections import defaultdict
from datetime import date
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from .forms import (
    ArabicAuthenticationForm,
    RegisterForm,
    BatchForm,
    SessionForm,
    TraineeAdminForm,
    TraineeImportForm,
    ProfileForm,
    SiteSettingForm,
)
from .models import (
    Batch,
    Session,
    TraineeProfile,
    AttendanceRecord,
    Announcement,
    TraineeZoomLink,
    SiteSetting,
    Notification,
)


def is_admin(user):
    return user.is_authenticated and user.is_staff


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def get_profile(user):
    profile, _ = TraineeProfile.objects.get_or_create(
        user=user,
        defaults={'full_name': user.get_full_name() or user.username}
    )
    return profile


def home(request):
    if not request.user.is_authenticated:
        return redirect('auth')
    if request.user.is_staff:
        return redirect('admin_overview')
    return redirect('trainee_home')


def auth_view(request):
    if request.user.is_authenticated:
        return redirect('home')

    active_tab = request.POST.get('action', 'login')
    login_form = ArabicAuthenticationForm(request, data=request.POST if request.POST.get('action') == 'login' else None)
    register_form = RegisterForm(request.POST if request.POST.get('action') == 'register' else None)

    if request.method == 'POST' and request.POST.get('action') == 'login':
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password') or ''
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            login(request, user)
            messages.success(request, f'أهلًا بك، {user.first_name or user.username}')
            return redirect('home')
        messages.error(request, 'بيانات الدخول غير صحيحة أو الحساب غير مفعّل.')

    if request.method == 'POST' and request.POST.get('action') == 'register':
        if register_form.is_valid():
            user = User.objects.create_user(
                username=register_form.cleaned_data['username'],
                password=register_form.cleaned_data['password'],
                first_name=register_form.cleaned_data['full_name'],
                email=register_form.cleaned_data.get('email') or '',
            )
            TraineeProfile.objects.create(
                user=user,
                full_name=register_form.cleaned_data['full_name'],
                phone=register_form.cleaned_data.get('phone') or '',
                batch=register_form.cleaned_data.get('batch'),
            )
            login(request, user)
            messages.success(request, 'تم إنشاء حسابك بنجاح، أهلًا بك في أكاديميتي')
            return redirect('trainee_home')
        messages.error(request, 'توجد أخطاء في نموذج إنشاء الحساب.')

    return render(request, 'attendance/auth.html', {
        'login_form': login_form,
        'register_form': register_form,
        'active_tab': active_tab,
    })


@login_required
@user_passes_test(is_admin)
def admin_overview(request):
    today = timezone.localdate()
    trainees_count = TraineeProfile.objects.count()
    batches_count = Batch.objects.filter(is_active=True).count()
    sessions_today = Session.objects.filter(date=today, is_active=True).select_related('batch')
    today_records = AttendanceRecord.objects.filter(session__date=today, status='present')
    sessions_total = Session.objects.count()
    possible_today = sum(s.batch.trainees.count() for s in sessions_today)
    today_rate = round((today_records.count() / possible_today) * 100) if possible_today else 0
    recent_attendance = AttendanceRecord.objects.select_related('trainee', 'session')[:8]
    batch_stats = Batch.objects.filter(is_active=True)[:5]
    announcements = Announcement.objects.filter(is_active=True).filter(Q(audience='all') | Q(audience='admin'))[:3]
    return render(request, 'attendance/admin_overview.html', {
        'today': today,
        'trainees_count': trainees_count,
        'batches_count': batches_count,
        'sessions_today_count': sessions_today.count(),
        'today_rate': today_rate,
        'recent_attendance': recent_attendance,
        'batch_stats': batch_stats,
        'sessions_total': sessions_total,
        'announcements': announcements,
    })


@login_required
@user_passes_test(is_admin)
def admin_batches(request):
    edit_id = request.GET.get('edit')
    instance = get_object_or_404(Batch, pk=edit_id) if edit_id else None
    form = BatchForm(instance=instance)
    if request.method == 'POST':
        instance = get_object_or_404(Batch, pk=request.POST.get('batch_id')) if request.POST.get('batch_id') else None
        form = BatchForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم حفظ الدفعة بنجاح')
            return redirect('admin_batches')
        messages.error(request, 'لم يتم حفظ الدفعة؛ راجعي الحقول المطلوبة.')
    batches = Batch.objects.all().prefetch_related('trainees', 'sessions')
    return render(request, 'attendance/admin_batches.html', {'batches': batches, 'form': form, 'edit_instance': instance})


@login_required
@user_passes_test(is_admin)
def delete_batch(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    if request.method == 'POST':
        batch.delete()
        messages.success(request, 'تم حذف الدفعة')
    return redirect('admin_batches')


@login_required
@user_passes_test(is_admin)
def admin_sessions(request):
    edit_id = request.GET.get('edit')
    instance = get_object_or_404(Session, pk=edit_id) if edit_id else None
    form = SessionForm(instance=instance)
    if request.method == 'POST':
        instance = get_object_or_404(Session, pk=request.POST.get('session_id')) if request.POST.get('session_id') else None
        form = SessionForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم حفظ المحاضرة ورابط الزووم')
            return redirect('admin_sessions')
        messages.error(request, 'لم يتم حفظ المحاضرة؛ تأكدي من الرابط والتاريخ.')
    batch_id = request.GET.get('batch')
    sessions = Session.objects.select_related('batch')
    if batch_id:
        sessions = sessions.filter(batch_id=batch_id)
    sessions = sessions[:250]
    return render(request, 'attendance/admin_sessions.html', {
        'sessions': sessions,
        'form': form,
        'batches': Batch.objects.filter(is_active=True),
        'selected_batch': batch_id,
        'today': timezone.localdate(),
        'edit_instance': instance,
    })


@login_required
@user_passes_test(is_admin)
def delete_session(request, pk):
    session = get_object_or_404(Session, pk=pk)
    if request.method == 'POST':
        session.delete()
        messages.success(request, 'تم حذف المحاضرة')
    return redirect('admin_sessions')


def normalize_header(value):
    return str(value or '').strip().lower().replace(' ', '_').replace('-', '_')


def get_cell(row_map, *names):
    for name in names:
        key = normalize_header(name)
        value = row_map.get(key)
        if value not in (None, ''):
            return str(value).strip()
    return ''


def import_trainees_from_excel(file_obj, default_batch=None, update_existing=True, default_password=''):
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {'created': 0, 'updated': 0, 'skipped': 0, 'errors': ['الملف فارغ.']}

    headers = [normalize_header(h) for h in rows[0]]
    header_has_names = any(h in headers for h in ['username', 'اسم_المستخدم', 'full_name', 'الاسم', 'الاسم_الكامل'])
    data_rows = rows[1:] if header_has_names else rows
    if not header_has_names:
        headers = ['full_name', 'username', 'phone', 'email', 'batch', 'password']

    created = updated = skipped = 0
    errors = []
    batch_cache = {b.name.strip(): b for b in Batch.objects.all()}

    with transaction.atomic():
        for idx, row in enumerate(data_rows, start=2 if header_has_names else 1):
            row_map = {headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
            full_name = get_cell(row_map, 'full_name', 'name', 'الاسم', 'الاسم الكامل', 'اسم المتدرب')
            username = get_cell(row_map, 'username', 'user', 'اسم المستخدم')
            phone = get_cell(row_map, 'phone', 'mobile', 'رقم الجوال', 'الجوال')
            email = get_cell(row_map, 'email', 'البريد الإلكتروني', 'البريد')
            batch_name = get_cell(row_map, 'batch', 'group', 'دفعة', 'الدفعة')
            password = get_cell(row_map, 'password', 'كلمة المرور', 'كلمة_المرور') or default_password

            if not full_name and not username:
                continue
            if not username:
                skipped += 1
                errors.append(f'الصف {idx}: اسم المستخدم مطلوب.')
                continue
            if not full_name:
                full_name = username
            if not password:
                skipped += 1
                errors.append(f'الصف {idx}: كلمة المرور مطلوبة أو ضعي كلمة مرور موحدة في نموذج الاستيراد.')
                continue

            selected_batch = default_batch
            if batch_name:
                selected_batch = batch_cache.get(batch_name)
                if selected_batch is None:
                    selected_batch = Batch.objects.create(name=batch_name, description='أُنشئت تلقائيًا من ملف Excel')
                    batch_cache[batch_name] = selected_batch

            user = User.objects.filter(username=username).first()
            if user and not update_existing:
                skipped += 1
                continue
            if user:
                user.first_name = full_name
                user.email = email
                user.is_staff = False
                user.is_superuser = False
                user.set_password(password)
                user.save(update_fields=['first_name', 'email', 'is_staff', 'is_superuser', 'password'])
                profile, _ = TraineeProfile.objects.get_or_create(user=user)
                updated += 1
            else:
                user = User.objects.create(
                    username=username,
                    first_name=full_name,
                    email=email,
                    password=make_password(password),
                    is_staff=False,
                    is_superuser=False,
                )
                profile = TraineeProfile(user=user)
                created += 1

            profile.full_name = full_name
            profile.phone = phone
            profile.batch = selected_batch
            profile.save()

    return {'created': created, 'updated': updated, 'skipped': skipped, 'errors': errors[:20]}


@login_required
@user_passes_test(is_admin)
def admin_trainees(request):
    edit_id = request.GET.get('edit')
    instance = get_object_or_404(TraineeProfile, pk=edit_id) if edit_id else None
    form = TraineeAdminForm(instance=instance)
    import_form = TraineeImportForm()

    if request.method == 'POST' and request.POST.get('action') == 'manual':
        instance = get_object_or_404(TraineeProfile, pk=request.POST.get('trainee_id')) if request.POST.get('trainee_id') else None
        form = TraineeAdminForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم حفظ بيانات المتدرب')
            return redirect('admin_trainees')
        messages.error(request, 'لم يتم حفظ بيانات المتدرب؛ راجعي الحقول المطلوبة.')

    if request.method == 'POST' and request.POST.get('action') == 'import':
        import_form = TraineeImportForm(request.POST, request.FILES)
        if import_form.is_valid():
            result = import_trainees_from_excel(
                request.FILES['excel_file'],
                default_batch=import_form.cleaned_data.get('default_batch'),
                update_existing=import_form.cleaned_data.get('update_existing'),
                default_password=import_form.cleaned_data.get('default_password') or '',
            )
            msg = f"تم الاستيراد: {result['created']} جديد، {result['updated']} تحديث، {result['skipped']} متجاوز."
            messages.success(request, msg)
            for err in result['errors']:
                messages.error(request, err)
            return redirect('admin_trainees')
        messages.error(request, 'لم يتم استيراد الملف؛ تأكدي من صيغة Excel والحقول.')

    q = request.GET.get('q', '')
    batch_id = request.GET.get('batch', '')
    trainees = TraineeProfile.objects.select_related('user', 'batch')
    if q:
        trainees = trainees.filter(Q(full_name__icontains=q) | Q(phone__icontains=q) | Q(user__username__icontains=q) | Q(user__email__icontains=q))
    if batch_id:
        trainees = trainees.filter(batch_id=batch_id)
    paginator = Paginator(trainees, 100)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'attendance/admin_trainees.html', {
        'trainees': page_obj.object_list,
        'page_obj': page_obj,
        'batches': Batch.objects.filter(is_active=True),
        'form': form,
        'import_form': import_form,
        'q': q,
        'selected_batch': batch_id,
        'edit_instance': instance,
    })


@login_required
@user_passes_test(is_admin)
def export_trainees_csv(request):
    q = request.GET.get('q', '')
    batch_id = request.GET.get('batch', '')
    trainees = TraineeProfile.objects.select_related('user', 'batch')
    if q:
        trainees = trainees.filter(Q(full_name__icontains=q) | Q(phone__icontains=q) | Q(user__username__icontains=q) | Q(user__email__icontains=q))
    if batch_id:
        trainees = trainees.filter(batch_id=batch_id)
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="academy_trainees.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['الاسم الكامل', 'اسم المستخدم', 'رقم الجوال', 'البريد الإلكتروني', 'الدفعة', 'نسبة الحضور', 'تاريخ التسجيل'])
    for t in trainees:
        writer.writerow([t.full_name, t.user.username, t.phone, t.user.email, t.batch.name if t.batch else '', f'{t.attendance_rate}%', timezone.localtime(t.created_at).strftime('%Y-%m-%d %H:%M')])
    return response


@login_required
@user_passes_test(is_admin)
def delete_trainee(request, pk):
    profile = get_object_or_404(TraineeProfile, pk=pk)
    if request.method == 'POST':
        user = profile.user
        user.delete()
        messages.success(request, 'تم حذف المتدرب')
    return redirect('admin_trainees')


def filtered_attendance_queryset(request):
    q = request.GET.get('q', '')
    batch_id = request.GET.get('batch', '')
    status = request.GET.get('status', '')
    today = timezone.localdate()
    date_from = request.GET.get('date_from') or today.isoformat()
    date_to = request.GET.get('date_to') or today.isoformat()

    sessions = Session.objects.select_related('batch').filter(is_active=True)
    trainees = TraineeProfile.objects.select_related('user', 'batch')

    if q:
        trainees = trainees.filter(Q(full_name__icontains=q) | Q(phone__icontains=q) | Q(user__username__icontains=q) | Q(user__email__icontains=q))
    if batch_id:
        sessions = sessions.filter(batch_id=batch_id)
        trainees = trainees.filter(batch_id=batch_id)
    if date_from:
        sessions = sessions.filter(date__gte=date_from)
    if date_to:
        sessions = sessions.filter(date__lte=date_to)

    sessions = list(sessions.order_by('-date', '-start_time')[:60])
    trainees = list(trainees)
    trainees_by_batch = defaultdict(list)
    for trainee in trainees:
        if trainee.batch_id:
            trainees_by_batch[trainee.batch_id].append(trainee)

    session_ids = [s.id for s in sessions]
    trainee_ids = [t.id for t in trainees]
    records = AttendanceRecord.objects.filter(session_id__in=session_ids, trainee_id__in=trainee_ids).select_related('trainee', 'session')
    record_map = {(r.trainee_id, r.session_id): r for r in records}

    rows = []
    for session in sessions:
        for trainee in trainees_by_batch.get(session.batch_id, []):
            record = record_map.get((trainee.id, session.id))
            computed_status = record.status if record else 'absent'
            if status and computed_status != status:
                continue
            rows.append({'session': session, 'trainee': trainee, 'record': record, 'status': computed_status})
    return rows, {'q': q, 'batch': batch_id, 'status': status, 'date_from': date_from, 'date_to': date_to}


def send_absence_notice(trainee, session, sender_user=None):
    title = 'تنبيه غياب عن محاضرة'
    body = f'نأمل الانتباه: لم يتم تسجيل حضورك في محاضرة "{session.title}" بتاريخ {session.date}. يرجى مراجعة الإدارة عند وجود عذر أو مشكلة تقنية.'
    Notification.objects.create(recipient=trainee.user, title=title, body=body, session=session)
    if trainee.user.email:
        try:
            send_mail(
                subject=title,
                message=body,
                from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
                recipient_list=[trainee.user.email],
                fail_silently=True,
            )
        except Exception:
            pass


@login_required
@user_passes_test(is_admin)
def admin_attendance(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'send_one':
            trainee = get_object_or_404(TraineeProfile, pk=request.POST.get('trainee_id'))
            session = get_object_or_404(Session, pk=request.POST.get('session_id'))
            if not AttendanceRecord.objects.filter(trainee=trainee, session=session, status='present').exists():
                send_absence_notice(trainee, session, request.user)
                messages.success(request, f'تم إرسال تنبيه الغياب إلى {trainee.full_name}')
            return redirect(request.get_full_path())
        if action == 'send_all_absent':
            rows, _ = filtered_attendance_queryset(request)
            sent = 0
            for row in rows:
                if row['status'] == 'absent':
                    send_absence_notice(row['trainee'], row['session'], request.user)
                    sent += 1
            messages.success(request, f'تم إرسال تنبيهات الغياب إلى {sent} متدرب/ـة.')
            return redirect(request.get_full_path())

    rows, filters = filtered_attendance_queryset(request)
    present_count = sum(1 for r in rows if r['status'] == 'present')
    absent_count = sum(1 for r in rows if r['status'] == 'absent')
    paginator = Paginator(rows, 120)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'attendance/admin_attendance.html', {
        'rows': page_obj.object_list,
        'page_obj': page_obj,
        'filters': filters,
        'batches': Batch.objects.filter(is_active=True),
        'present_count': present_count,
        'absent_count': absent_count,
        'total_count': len(rows),
    })


@login_required
@user_passes_test(is_admin)
def export_attendance_csv(request):
    rows, _ = filtered_attendance_queryset(request)
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="academy_attendance.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['المتدرب', 'اسم المستخدم', 'الجوال', 'البريد', 'الدفعة', 'المحاضرة', 'التاريخ', 'الحالة', 'وقت التحضير', 'رابط زووم'])
    for row in rows:
        record = row['record']
        writer.writerow([
            row['trainee'].full_name,
            row['trainee'].user.username,
            row['trainee'].phone,
            row['trainee'].user.email,
            row['session'].batch.name,
            row['session'].title,
            row['session'].date,
            'حاضر' if row['status'] == 'present' else 'متأخر' if row['status'] == 'late' else 'غائب',
            timezone.localtime(record.checked_at).strftime('%Y-%m-%d %H:%M') if record else '-',
            record.zoom_url_snapshot if record else row['session'].zoom_url,
        ])
    return response


@login_required
@user_passes_test(is_admin)
def admin_reports(request):
    batch_id = request.GET.get('batch')
    batches = Batch.objects.filter(is_active=True)
    selected_batch = batches.filter(pk=batch_id).first() if batch_id else batches.first()
    trainees = TraineeProfile.objects.filter(batch=selected_batch).select_related('user') if selected_batch else []
    total_rate = selected_batch.attendance_rate if selected_batch else 0
    top = sorted(list(trainees), key=lambda p: p.attendance_rate, reverse=True)[:10]
    low = [p for p in trainees if p.attendance_rate < 60]
    return render(request, 'attendance/admin_reports.html', {
        'batches': batches,
        'selected_batch': selected_batch,
        'total_rate': total_rate,
        'top': top,
        'low': low,
    })


@login_required
@user_passes_test(is_admin)
def admin_settings(request):
    setting = SiteSetting.current()
    form = SiteSettingForm(instance=setting)
    if request.method == 'POST':
        form = SiteSettingForm(request.POST, instance=setting)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم حفظ إعدادات النظام')
            return redirect('admin_settings')
    return render(request, 'attendance/admin_settings.html', {'form': form})


@login_required
def trainee_home(request):
    if request.user.is_staff:
        return redirect('admin_overview')
    profile = get_profile(request.user)
    today = timezone.localdate()

    todays_sessions = []
    today_session_rows = []
    if profile.batch:
        todays_sessions = list(
            Session.objects.filter(batch=profile.batch, date=today, is_active=True)
            .order_by('start_time', 'id')
        )

    if todays_sessions:
        session_ids = [session.id for session in todays_sessions]
        records = AttendanceRecord.objects.filter(trainee=profile, session_id__in=session_ids)
        record_map = {record.session_id: record for record in records}
        custom_links = TraineeZoomLink.objects.filter(trainee=profile, session_id__in=session_ids)
        custom_link_map = {link.session_id: link.zoom_url for link in custom_links}

        for session in todays_sessions:
            record = record_map.get(session.id)
            zoom_url = custom_link_map.get(session.id) or session.zoom_url
            today_session_rows.append({
                'session': session,
                'record': record,
                'zoom_url': zoom_url if record else '',
                'raw_zoom_url': zoom_url,
            })

    announcements = Announcement.objects.filter(is_active=True).filter(Q(audience='all') | Q(audience='trainee'))[:3]
    notifications = Notification.objects.filter(recipient=request.user)[:6]
    total_sessions = profile.batch.sessions.filter(is_active=True).count() if profile.batch else 0
    attended_today_count = sum(1 for row in today_session_rows if row['record'])

    return render(request, 'attendance/trainee_home.html', {
        'profile': profile,
        'today': today,
        'todays_sessions': todays_sessions,
        'today_session_rows': today_session_rows,
        'attended_today_count': attended_today_count,
        'total_sessions': total_sessions,
        'announcements': announcements,
        'notifications': notifications,
    })


@login_required
def check_in(request, session_id):
    if request.user.is_staff:
        return redirect('admin_overview')
    profile = get_profile(request.user)
    session = get_object_or_404(Session, pk=session_id, batch=profile.batch, is_active=True)
    custom = TraineeZoomLink.objects.filter(session=session, trainee=profile).first()
    zoom_url = custom.zoom_url if custom else session.zoom_url
    AttendanceRecord.objects.get_or_create(
        trainee=profile,
        session=session,
        defaults={
            'status': 'present',
            'zoom_url_snapshot': zoom_url,
            'ip_address': get_client_ip(request),
            'user_agent': request.META.get('HTTP_USER_AGENT', '')[:1000],
        }
    )
    messages.success(request, 'تم تسجيل حضورك بنجاح')
    return redirect('trainee_home')


@login_required
def trainee_schedule(request):
    if request.user.is_staff:
        return redirect('admin_overview')
    profile = get_profile(request.user)
    sessions = list(Session.objects.filter(batch=profile.batch).order_by('-date', '-start_time', '-id')) if profile.batch else []
    attended_ids = set(AttendanceRecord.objects.filter(trainee=profile).values_list('session_id', flat=True))
    today = timezone.localdate()
    return render(request, 'attendance/trainee_schedule.html', {
        'profile': profile,
        'sessions': sessions,
        'attended_ids': attended_ids,
        'today': today,
    })


@login_required
def trainee_calendar(request):
    if request.user.is_staff:
        return redirect('admin_overview')
    profile = get_profile(request.user)
    today = timezone.localdate()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    cal = calendar.Calendar(firstweekday=6)
    weeks = cal.monthdatescalendar(year, month)
    records = AttendanceRecord.objects.filter(trainee=profile, session__date__year=year, session__date__month=month).select_related('session')
    present_days = {r.session.date.day: r for r in records}
    prev_month = month - 1 or 12
    prev_year = year - 1 if month == 1 else year
    next_month = month + 1 if month < 12 else 1
    next_year = year + 1 if month == 12 else year
    month_names = ['يناير', 'فبراير', 'مارس', 'أبريل', 'مايو', 'يونيو', 'يوليو', 'أغسطس', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']
    return render(request, 'attendance/trainee_calendar.html', {
        'profile': profile,
        'weeks': weeks,
        'month': month,
        'year': year,
        'month_name': month_names[month - 1],
        'present_days': present_days,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
    })


@login_required
def trainee_profile(request):
    if request.user.is_staff:
        return redirect('admin_overview')
    profile = get_profile(request.user)
    profile_form = ProfileForm(instance=profile)
    password_form = PasswordChangeForm(request.user)
    if request.method == 'POST' and request.POST.get('form_type') == 'profile':
        profile_form = ProfileForm(request.POST, instance=profile)
        if profile_form.is_valid():
            profile_form.save()
            messages.success(request, 'تم حفظ بياناتك الشخصية')
            return redirect('trainee_profile')
    if request.method == 'POST' and request.POST.get('form_type') == 'password':
        password_form = PasswordChangeForm(request.user, request.POST)
        if password_form.is_valid():
            user = password_form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'تم تحديث كلمة المرور')
            return redirect('trainee_profile')
    return render(request, 'attendance/trainee_profile.html', {
        'profile': profile,
        'profile_form': profile_form,
        'password_form': password_form,
    })
